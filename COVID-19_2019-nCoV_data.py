from base64 import encode
import encodings
from tkinter import N
from urllib.error import HTTPError
from wsgiref import headers
import urllib3
import requests
import time
from bs4 import BeautifulSoup
import openpyxl
import re

addresslist = ["上大路1288弄", "杨泰路99弄", "教育路555弄"]
updateurl = "https://chenfan.info/update_date/"
url = "https://chenfan.info/search_data/"
url1 = "https://chenfan.info/search_data/1=1||1=1"
url2 = "https://chenfan.info/search_data/上大路"
pageFile = "./DataFile/recorder.html"
xlsxFile = "./DataFile/recorder.xlsx"

dateRex = r"(\d{4}-\d{1,2}-\d{1,2})"


def getLevel(strCellId):
    idArray = str.split(strCellId, "_")
    return idArray[2]


urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
print("\033c")
print("开始检索")
t1 = time.time()

# temphtml="<th id=\"T_df6ca_level0_row0\" class=\"row_heading level0 row0\" rowspan=\"18\">宝山区</th>"
# soup = BeautifulSoup(temphtml, 'lxml')
# datacell = soup.find_all("th")[0]
# print(datacell)
# tt=BeautifulSoup(str(datacell),"lxml").th["id"]
# print(tt)
heads = {
    "accept":
    "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
    "accept-encoding":
    "gzip, deflate, br",
    "accept-language":
    "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6,ja;q=0.5,zh-TW;q=0.4,es;q=0.3,sl;q=0.2,ru;q=0.1,mt;q=0.1",
    "user-agent":
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.75 Safari/537.36 Edg/100.0.1185.36"
}

try:
    res = requests.get(updateurl, headers=heads)
    res.encoding = "utf-8"
    print("数据更新至:",res.text)
    res.close()
    res = requests.get(url2, headers=heads)
    res.encoding = "utf-8"
    print(res.text)
    lsitData = res.text.split("\n")
    res.close()

    # dataBook = openpyxl.Workbook()
    # dataSheet = dataBook.active
    # dataSheet.title = "上海感染记录"
    # dataSheet.cell(1, 1, "区")
    # dataSheet.cell(1, 2, "地址")
    # dataSheet.cell(1, 3, "感染记录日期")

    # soup = BeautifulSoup(res.text, 'lxml')
    # print(res.text)
    # res.close()
    # data_body = soup.find_all("tbody")
    # if (data_body == None or (len(data_body)==0)):
    #     warn=soup.find_all("h3")
    #     tips=soup.find_all("br")
    #     for w in warn:
    #         print(w.text)
    #     for tip in tips:
    #         print(tip.text)
    #     exit()
    # dataBook = openpyxl.Workbook()
    # dataSheet = dataBook.active
    # dataSheet.title = "上海感染记录"
    # dataSheet.cell(1, 1, "区")
    # dataSheet.cell(1, 2, "地址")
    # dataSheet.cell(1, 3, "解除剩余")
    # dataSheet.cell(1, 4, "感染记录日期")
    # cellList = data_body[0].find_all("th")
    # max = len(cellList)
    # area=cellList[0].string
    # i = 0
    # j = 1
    # while (i < max):
    #     thSoup = BeautifulSoup(str(cellList[i]), 'lxml')
    #     level = getLevel(thSoup.th["id"])
    #     if (level == "level3"):
    #         print(cellList[i].string, end="    ")
    #         if (dataSheet.cell(j, 4).value is None):
    #             dataSheet.cell(j, 4, cellList[i].string + ",")
    #         else:
    #             dataSheet.cell(
    #                 j, 4,
    #                 dataSheet.cell(j, 4).value + cellList[i].string + ",")
    #     else:
    #         if (level == "level0"):
    #             print()
    #             j = j + 1
    #             area=cellList[i].string
    #             print(cellList[i].string, end="    ")
    #             dataSheet.cell(j, 1, cellList[i].string)
    #         if (level == "level1"):
    #             prevSoup = BeautifulSoup(str(cellList[i-1]), 'lxml')
    #             prevlevel = getLevel(prevSoup.th["id"])
    #             if(prevlevel!="level0"):
    #                 print()
    #                 j = j + 1
    #                 print(area, end="    ")
    #             print(cellList[i].string, end="    ")
    #             dataSheet.cell(j, 1, area)
    #             dataSheet.cell(j, 2, cellList[i].string)
    #         if (level == "level2"):
    #             print(cellList[i].text, end="    ")
    #             dataSheet.cell(j, 3, cellList[i].text)
    #     i = i + 1

    # 老版本数据抓去逻辑
    # dataSheet.cell(2, 1, cellList[0].string)
    # dataSheet.cell(2, 2, cellList[1].string)
    # print(cellList[0].string, end="    ")
    # print(cellList[1].string, end="    ")
    # area = cellList[0].string
    # i = 2
    # j = 2
    # while (i < max):
    #     if (re.match(dateRex, cellList[i].string) is not None):
    #         print(cellList[i].string, end="    ")
    #         if (dataSheet.cell(j, 3).value is None):
    #             dataSheet.cell(j, 3, cellList[i].string + ",")
    #         else:
    #             dataSheet.cell(
    #                 j, 3,
    #                 dataSheet.cell(j, 3).value + cellList[i].string + ",")
    #         i = i + 1
    #     else:
    #         print()
    #         j = j + 1
    #         if (re.match(dateRex, cellList[i + 1].string) is not None):
    #             dataSheet.cell(j, 1, area)
    #             dataSheet.cell(j, 2, cellList[i].string)

    #             print(area, end="    ")
    #             print(cellList[i].string, end="    ")
    #             i = i + 1
    #         else:
    #             dataSheet.cell(j, 1, cellList[i].string)
    #             dataSheet.cell(j, 2, cellList[i + 1].string)

    #             area = cellList[i].string
    #             print(cellList[i].string, end="    ")
    #             print(cellList[i + 1].string, end="    ")
    #             i = i + 2
    # dataBook.save(filename=xlsxFile)
except HTTPError as httperr:
    print(httperr)
except OSError as err:
    print("访问地址出现异常，异常信息：", err)

print()
t2 = time.time()
print("数据检索完毕，全程耗时:", (t2 - t1) * 1000, "毫秒")