from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Alignment
import pandas as pd
import time
import sys
import re
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows

csvFile = r"./TASK_RTC/Task查询.csv"  #RTC平台下载的原文件名
newFile = r"./TASK_RTC/tmpfile.xlsx"  #临时中转文件
rtcFile = r"./TASK_RTC/VCS1.0_2.0+  WBS.xlsx"  #飞书表格名称
saveFile = r"./TASK_RTC/TASK_new.xlsx"  #最终合并文件名

extendColcolor = "DEE0E1"
rtcCorlor = "81817F"
devCorlor = "F28A15"
testCorlor = "17C1FF"
extName = [
    '车型', '类型', '模块', '责任人', 'PM', '内部状态', '第一轮可测试时间', '冻结时间', '发布时间', '超期/天',
    '接口变更', '提交代码日期', '依赖', '进度', '开发效果', '测试人员', '验证版本', '验证结果', 'Buglist',
    '签入日期', '备注'
]
moduleOnwers = {"Climate": "徐欢", "setting": "苗棽",
                "vehicle_control": "苗棽","adas_setting": "翟嘉兴", "drive_info": "苗棽",
                "vehicle_info": "苗棽", "ambient_light": "张书俊", "drive_mode": "翟家兴","adas_normal":"苗棽",
                "kanziservice_motion": "韩闯", "kanziservice": "刘波", "apa": "张书俊","arhud": "翟家兴",
                "Cluster_ADAS":"张书俊", "cluster_alert": "翟嘉兴", "cluster_warning": "翟嘉兴", "Drive_Mode": "卢伟国",
                "hud": "苗棽", "car_model": "张书俊", "performance": "金正轩", "framework": "翟嘉兴","launcher": "翟嘉兴",
                "peekin": "卢伟国","charging": "卢伟国", "Vehicle Setting": "徐欢","cabinmode": "苗棽","lvm_lpnp":"韩闯"}

# 主进程开始---------------------------------
if __name__ == '__main__':
    name = sys.argv[1:2]
    name = ['RTC TASK']
    if (len(name) < 1 or len(name[0]) < 1):
        print('RTC sheet name is empty ')
        sys.exit()
    else:
        # rtcSheetName = name[0]
        print('name is ', name[0])

sTime = time.time()
print("CSV文件转存xlsx文件......")
# 创建一个workbook
xlsbook = openpyxl.Workbook()
# 创建一个worksheet
newsheet = xlsbook.create_sheet('RTC TASK', -1)
# newsheet.title='Open assigned to me'

with open(csvFile, 'r', encoding='utf-16') as csv:
    row = csv.read().split('"\n')
    for line in row:
        line = re.sub('"', '', line)
        cells = line.split('	')
        newsheet.append(cells)

extCol = newsheet.max_column + 1
align = Alignment(horizontal='center', vertical='center')
extColorFill = PatternFill(start_color=extendColcolor, fill_type="solid")
devCorlorFill = PatternFill(start_color=devCorlor, fill_type="solid")
testCorlorFill = PatternFill(start_color=testCorlor, fill_type="solid")
for i in range(0, len(extName)):
    newsheet.cell(row=1, column=extCol + i).value = extName[i]
    newsheet.cell(row=1, column=extCol + i).alignment = align
    if (i < 10):
        newsheet.cell(row=1, column=extCol + i).fill = extColorFill
    if (9 < i < 15):
        newsheet.cell(row=1,column=extCol + i).fill = devCorlorFill
    if (14 < i < 19):
        newsheet.cell(row=1,column=extCol + i).fill = testCorlorFill

xlsbook.save(newFile)

print("文件转存完毕......")


print("读取表格文件数据")
taskSheet = pd.read_excel(newFile, sheet_name='RTC TASK')
rtcSheet = pd.read_excel(rtcFile, sheet_name='RTC TASK')

print("新增及已存在bug记录处理")

# 将taskSheet和rtcSheet设置为a_id和b_id为索引
taskSheet.set_index('Id', inplace=True)
rtcSheet.set_index('Id', inplace=True)
taskSheet['签入日期'] = taskSheet['签入日期'].apply(pd.to_datetime)

# 使用merge函数合并两个数据表，on参数设置为'Id'，表示以'Id'列作为主键进行合并
merged = pd.merge(taskSheet, rtcSheet, on='Id')

if not rtcSheet.index.is_unique:
    print("rtcSheet索引不唯一，存在重复值：")
    print(rtcSheet.index[rtcSheet.index.duplicated()])
    sys.exit()

if not taskSheet.index.is_unique:
    print("newSheet索引不唯一，存在重复值：")
    print(taskSheet.index[taskSheet.index.duplicated()])
    sys.exit()

# 使用update函数更新taskSheet中的记录
taskSheet.update(rtcSheet[extName])

# 补充签入日期
taskSheet.loc[taskSheet['签入日期'].isnull(), '签入日期'] = datetime.now().strftime("%Y-%m-%d")

taskbook = openpyxl.Workbook()
# 创建一个worksheet
savesheet = taskbook.create_sheet('RTC TASK', -1)

# 写入表头（包含索引列Id）
savesheet.append(["ID"] + taskSheet.columns.tolist())  

# 写入数据行（包含索引列Id）
for idx, row in taskSheet.iterrows():
    savesheet.append([idx] + row.tolist())



date_style = NamedStyle(name="date_style")
date_style.number_format = 'YYYY-MM-DD'
for row in savesheet.iter_rows(min_col=33, max_col=33):
    for cell in row:
        cell.style = date_style

for i in range(0, len(extName)):
    # newsheet.cell(row=1, column=extCol + i).value = extName[i]
    # newsheet.cell(row=1, column=extCol + i).alignment = align
    if (i < 10):
        savesheet.cell(row=1, column=extCol + i).fill = extColorFill
    if (9 < i < 15):
        savesheet.cell(row=1,column=extCol + i).fill = devCorlorFill
    if (14 < i < 19):
        savesheet.cell(row=1,column=extCol + i).fill = testCorlorFill



# 保存数据到表格
taskbook.save(saveFile)



eTime = time.time()
print("数据处理完成，全程耗时：", str((eTime - sTime) * 1000), "毫秒。")
