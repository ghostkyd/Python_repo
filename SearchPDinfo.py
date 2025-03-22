from operator import mod
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import time
import os
import threading
from threading import Lock, Thread
from multiprocessing import Process

dirPath = r'D:\tmp\户籍数据1.03GB公安专用'
txtPath = r'D:\tmp\txtLogs'
process_list = []


def creatProcess(targ, filePath, fileName):
    tt = Process(target=targ, args=(filePath, fileName))
    tt.start()
    process_list.append(tt)


def writeDataToTxt(filePath, val):
    with open(filePath, 'a+') as f:
        f.writelines(str(val))


def searchData(filePath, fileName):
    path = os.path.join(filePath, fileName)
    namearr = str(fileName).split('.')
    txtName = namearr[0]+'.txt'
    print('Read data from %s ......' % fileName)
    dataBook = load_workbook(path, read_only=True, data_only=True)
    namelist = dataBook.sheetnames
    dataSheet = dataBook[namelist[0]]
    print('%s data loaded, start to search......' % fileName)
    for i in range(1, dataSheet.max_row):
        name = str(dataSheet.cell(row=i, column=8).value).strip()
        if ('王启嘉' == name):
            for cell in dataSheet[i]:
                val = val+cell+','
            val = val[0:len(val)-1]
            writePath = os.path.join(txtPath, txtName)
            writeDataToTxt(writePath, val)
    dataBook.close()
    print('%s done......' % fileName)

# tt=r'E:\户籍数据1.03GB公安专用\1 - 副本 (10).xlsx'
# stime = time.time()
# searchData(tt)
# etime = time.time()
# print("Process time: %0.3f seconds..."%(etime - stime))


if __name__ == '__main__':
    os.chdir(dirPath)
    dirList = os.listdir()

    for ll in dirList:
        path = os.path.join(dirPath, ll)
        if (not os.path.isdir(path)):
            creatProcess(searchData, dirPath, ll)

    for i in process_list:
        i.join()

    print("主进程结束！")
