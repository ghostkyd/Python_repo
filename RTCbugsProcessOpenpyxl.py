from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Alignment
import pandas as pd
import time
import sys
import re

csvFile = r"./DataFile/RTC每日更新查询.csv"
newFile = r"./DataFile/Open assigned to me.xlsx"
rtcFile = r"./DataFile/RTC系统问题汇总--每天更新_2025.xlsx"
saveFile = r"./DataFile/RTC list new.xlsx"

platform_GB= ['B233', 'B223', 'E22', 'E2UB', 'E2YB', 'C1YB','L232','L233','L234','C1UL_2','O1SL_2','A2LL','E2UL','E2QL']
platform_CL = ['458', '458-HEV','358-2','358-2PHEV','557']

extendColcolor = "FDE9D9"
extName = ['领域', '说明', '模块', '车型', '负责人', '超期/天', '进展说明（对应模块的研发更新）', '研发状态开发更新',
           '提交代码date', '版本时间', '测试结果', '测试备注', '签入日期','测试人员']

moduleOnwers = {"climate": "徐欢", "audio": "张书俊", "setting": "苗棽",
                "vehicle_control": "苗棽","adas_setting": "翟嘉兴", "drive_info": "苗棽",
                "vehicle_info": "苗棽", "ambient_light": "张书俊", "drive_mode": "翟家兴","adas_normal":"苗棽",
                "kanziservice_motion": "韩闯", "kanziservice": "刘波", "apa": "张书俊","arhud": "翟家兴",
                "cluster_adas":"张书俊", "cluster_alert": "翟嘉兴", "cluster_warning": "翟嘉兴", "cluster_wallpaper": "卢伟国",
                "hud": "苗棽", "car_model": "张书俊", "performance": "金正轩", "framework": "翟嘉兴","launcher": "翟嘉兴",
                "peekin": "卢伟国","charging": "卢伟国", "vehicle_setting": "徐欢","cabinmode": "苗棽","lvm_lpnp":"韩闯"}


# 功能函数块开始---------------------------------------
def getModule(summary):
    if ("climate" in str(summary).lower()):
        return "climate"
    if ("vehicle control" in str(summary).lower() or "vehicle_control" in str(summary).lower() or "vehiclecontrol" in str(summary).lower()):
        return "vehicle_control"
    if ("vehicle setting" in str(summary).lower() or "vehicle_setting" in str(summary).lower() or "vehiclesetting" in str(summary).lower()):
        return "vehicle_setting"
    if ("adassetting" in str(summary).lower() or "adas_setting" in str(summary).lower() or "adas setting" in str(summary).lower() or
            "adassettings" in str(summary).lower() or "adas_settings" in str(summary).lower() or "adas settings" in str(summary).lower()):
        return "adas_setting"
    if ("vehicle info" in str(summary).lower() or "vehicle_info" in str(summary).lower() or "vehicleinfo" in str(summary).lower() or "vehicle_state" in str(summary).lower()):
        return "vehicle_info"
    if ("driveinfo" in str(summary).lower() or "drive_info" in str(summary).lower() or "drive info" in str(summary).lower()):
        return "drive_info"
    if ("ambientlight" in str(summary).lower() or "ambient_light" in str(summary).lower() or "ambient light" in str(summary).lower()):
        return "ambient_light"
    if ("drivemode" in str(summary).lower() or "drive_mode" in str(summary).lower() or "drive mode" in str(summary).lower()):
        return "drive_mode"
    if ("alert" in str(summary).lower()):
        return "cluster_alert"
    if ("adas" in str(summary).lower() and "setting" not in str(summary).lower() and "settings" not in str(summary).lower()):
        return "cluster_adas"
    if ("hud" in str(summary).lower()):
        return "hud"
    if ("seats" in str(summary).lower() or "seat" in str(summary).lower()):
        return "seats"
    if ("audio" in str(summary).lower() or "audio basic" in str(summary).lower() or "audio_basic" in str(summary).lower() or "audiobasic" in str(summary).lower()):
        return "audio"
    if (("setting" in str(summary).lower() or "settings" in str(summary).lower()) and "adassetting" not in str(summary).lower() and "adas_setting" not in str(summary).lower() and "adas setting" not in str(summary).lower() and
            "adassettings" not in str(summary).lower() and "adas_settings" not in str(summary).lower() and "adas settings" not in str(summary).lower()):
        return "setting"
    if ("basic_module" in str(summary).lower() or "basic module" in str(summary).lower() or "basicmodule" in str(summary).lower()):
        return "basic_module"
    if ("cluster_warning" in str(summary).lower()):
        return "cluster_warning"
    if ("cluster" in str(summary).lower()):
        return "cluster_adas"
    if ("[performance]" in str(summary).lower()):
        return "performance"
    if ("[Framework]" in str(summary).lower()):
        return "framework"
    if ("[Launcher]" in str(summary).lower()):
        return "launcher"
    if ("peekin" in str(summary).lower()):
        return "peekin"
    if ("peek in" in str(summary).lower()):
        return "peekin"
    if ("Energy" in str(summary).lower()):
        return "charging"
    if ("空调" in str(summary).lower()):
        return "climate"
    if ("cabinmode" in str(summary).lower() or "cabin mode" in str(summary).lower() or "seat" in str(summary).lower()):
        return 'cabinmode'
    if ("lvm" in str(summary).lower() or "lpnp" in str(summary).lower()):
        return 'lvm_lpnp'
    if ("557" in str(summary).lower() and "apa" in str(summary).lower()):
        return 'lvm_lpnp'
    else:
        return ""


def getCarType(platform):
    carType = ""
    if ("Epsilon/E2UB/" in str(platform)):
        carType = carType+"E2UB,"
    if ("Epsilon/E2YB/" in str(platform)):
        carType = carType+"E2YB,"
    if ("BEV 3/B223/" in str(platform)):
        carType = carType+"B223,"
    if ("BEV 3/B233/" in str(platform)):
        carType = carType+"B233,"
    if ("Crossover/C1YB-2/" in str(platform)):
        carType = carType+"C1YB,"
    if ("/E2LB-2/" in str(platform)):
        carType = carType+"E22,"
    if ("U-Van/458/MY23" in str(platform)):
        carType = carType+"458,"
    if ("U-Van/458/MY25" in str(platform)):
        carType = carType+"458-HEV,"
    if ("U-Van/458/MY25.5" in str(platform)):
        carType = carType+"458-HEV,"
    if ("U-Van/458/MY26" in str(platform)):
        carType = carType+"458-HEV,"
    if ("U-Van/458 HEV/" in str(platform)):
        carType = carType+"458-HEV,"
    if ("U-Van/358-2/" in str(platform)):
        carType = carType+"358-2,"
    if ("U-Van/358-2 PHEV/" in str(platform)):
        carType = carType+"358-2PHEV,"
    if ("/NDLB/" in str(platform)):
        carType = carType+"NDLB,"
    if ("U-Van/557/MY26" in str(platform)):
        carType = carType+"557,"
    
    if ("Alpha/A2LL/" in str(platform)):
        carType = carType+"A2LL,"
    if ("BEV 3/L234" in str(platform)):
        carType = carType+"L234,"
    if ("BEV 3/L233" in str(platform)):
        carType = carType+"L233,"
    if ("BEV 3/L232" in str(platform)):
        carType = carType+"L232,"
    if ("Omega/O1SL-2" in str(platform)):
        carType = carType+"O1SL-2,"
    if ("Epsilon/E2UL" in str(platform)):
        carType = carType+"E2UL,"

        
    carType = carType[:-1]
    return carType


def getOwner(carType, module):
    try:
        if (module is None or type is None):
            return ""
        else:
            return moduleOnwers[module]
    except KeyError:
        # print(module)
        return ""

# 功能函数块结束------------------------------


# 主进程开始---------------------------------
if __name__ == '__main__':
    name = sys.argv[1:2]
    # name = ['1月17日']
    if (len(name) < 1 or len(name[0]) < 1):
        print('RTC sheet name is empty ')
        sys.exit()
    else:
        rtcSheetName = name[0]
        print('name is ', rtcSheetName)

sTime = time.time()
print("CSV文件转存xlsx文件......")
# 创建一个workbook
xlsbook = openpyxl.Workbook()
# 创建一个worksheet
newsheet = xlsbook.create_sheet('Open assigned to me', -1)
# newsheet.title='Open assigned to me'

with open(csvFile, 'r', encoding='utf-16') as csv:
    row = csv.read().split('"\n')
    for line in row:
        line = re.sub('"', '', line)
        cells = line.split('	')
        newsheet.append(cells)

extCol = newsheet.max_column+1
align = Alignment(horizontal='center', vertical='center')
for i in range(0, len(extName)):
    newsheet.cell(row=1, column=extCol+i).value = extName[i]
    newsheet.cell(row=1, column=extCol+i).alignment = align
    newsheet.cell(row=1, column=extCol +
                  i).fill = PatternFill(start_color=extendColcolor, fill_type="solid")

xlsbook.save(newFile)
print("文件转存完毕......")

print("读取表格文件数据")
newSheet = pd.read_excel(newFile, sheet_name='Open assigned to me')
rtcSheet = pd.read_excel(rtcFile, sheet_name=name[0])

# 删除"Type"列不是"Bug"的记录
newSheet = newSheet[newSheet['Type'] == 'Bug']

print("新增及已存在bug记录处理")

# 将newSheet和rtcSheet设置为a_id和b_id为索引
newSheet.set_index('Id', inplace=True)
rtcSheet.set_index('Id', inplace=True)

# 使用merge函数合并两个数据表，on参数设置为'Id'，表示以'Id'列作为主键进行合并
merged = pd.merge(newSheet, rtcSheet, on='Id')

# 初始化一个空数组id_tmp
id_tmp = []

# 遍历合并后的数据表
for index, row in merged.iterrows():
    # 如果newSheet数据表中“tag”列含有"must"字符，而rtcSheet数据表中“tag”不列含有"must"字符
    if 'must' in str(row['Tags_x']) and 'must' not in str(row['Tags_y']):
        # 就把这条记录的主键值保存在数组id_tmp中
        id_tmp.append(index)

if not rtcSheet.index.is_unique:
    print("rtcSheet索引不唯一，存在重复值：")
    print(rtcSheet.index[rtcSheet.index.duplicated()])
    sys.exit()

if not newSheet.index.is_unique:
    print("newSheet索引不唯一，存在重复值：")
    print(newSheet.index[newSheet.index.duplicated()])
    sys.exit()

# 使用update函数更新newSheet中的记录
newSheet.update(rtcSheet[extName])

# 之前被标记已转出但依然挂着的记录标记为已打回
newSheet.loc[newSheet['研发状态开发更新']=='已转出', '研发状态开发更新'] = '已转回'

# 研发状态为空的进行初始化
newSheet.loc[newSheet['研发状态开发更新'].isnull(), '研发状态开发更新'] = '待分析'

# 补充签入日期
newSheet.loc[newSheet['签入日期'].isnull(), '签入日期'] = datetime.now()

# 计算超期天数
newSheet['Creation Date'] = newSheet['Creation Date'].str.replace('上午', '')
newSheet['Creation Date'] = newSheet['Creation Date'].str.replace('下午', '')
# newSheet['Due Date']=datetime.strptime(newSheet['Due Date'],'%Y-%m-%d %p%I:%M')
newSheet['Creation Date'] = pd.to_datetime(newSheet['Creation Date'])
newSheet.loc[(newSheet['研发状态开发更新'] != '已转出') & (newSheet['Creation Date'].notnull(
)), '超期/天'] = (datetime.now() - newSheet['Creation Date']).dt.days

# 使用concat函数将不存在于newSheet中的b_id记录追加到newSheet的末尾
print("追加已迁出记录")
rtcSheet.loc[~rtcSheet.index.isin(newSheet.index), '研发状态开发更新'] = '已转出'
newSheet = pd.concat([newSheet, rtcSheet[~rtcSheet.index.isin(newSheet.index)]])

# 补充bug所属模块
newSheet.loc[newSheet['模块'].isnull(), '模块'] = newSheet['Summary'].apply(getModule)
# newSheet['模块'] = newSheet['Summary'].apply(getModule)

# 补充bug所属车型
newSheet.loc[newSheet['车型'].isnull(), '车型'] = newSheet['Platform Vehicle Model'].apply(getCarType)
# newSheet['车型'] = newSheet['Platform Vehicle Model'].apply(getCarType)

# 补充bug所处Owner
newSheet.loc[newSheet['负责人'].isnull(), '负责人'] = newSheet.apply(
    lambda row: getOwner(row['车型'], row['模块']), axis=1)
# newSheet['负责人'] = newSheet.apply(lambda row: getOwner(row['车型'], row['模块']), axis=1)

newSheet.to_excel(saveFile)
print("数据处理完毕，开始调整单元格式......")
xl_book = openpyxl.load_workbook(saveFile, read_only=False, data_only=True)
xl_sheet = xl_book['Sheet1']

rmv_fill = PatternFill(fill_type="solid", fgColor="808080")
addmust_fill = PatternFill(fill_type="solid", fgColor="ED4197")

# 遍历C列
for row in xl_sheet.iter_rows():
    if (str(row[0]) in id_tmp):
        for c in row:
            c.fill = addmust_fill
    cell = row[20]
    # 如果C列的值是“已转出”，就把当前行的背景色设置为灰色
    if (cell.value == '已转出' or cell.value == '已转出'):
        for cell in row:
            cell.fill = rmv_fill

# for col in ['U', 'V', 'Y']:
for col in ['V', 'W', 'Z']:
    for cell in xl_sheet[col]:
        cell.number_format = 'YYYY"年"M"月"D"日"'

xl_book.save(saveFile)
xl_book.close()

eTime = time.time()
print("数据处理完成，全程耗时：", str((eTime-sTime)*1000), "毫秒。")
