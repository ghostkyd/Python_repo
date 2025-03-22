from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Alignment
import pandas as pd
import time
import sys
import re

csvFile = r"./QNX_RTC/RTC Bug汇总.csv"     #RTC平台下载的原文件名
newFile = r"./QNX_RTC/tmpfile.xlsx"             #临时中转文件
rtcFile = r"./QNX_RTC/RTC Bug汇总---仪表.xlsx"    #飞书表格名称
saveFile = r"./QNX_RTC/RTC list new.xlsx"        #最终合并文件


extendColcolor = "FDE9D9"
extName = ['分支车型','超期/天（Creation Date）','计划完成时间','进展说明（对应模块的研发更新）','研发状态开发更新',
           'R5分支提交date','主分支提交date','hotfix提交date','版本时间','测试结果','签入日期']



# 功能函数块开始---------------------------------------


def getPlatform(platStr):
    if("U-Van" in str(platStr)):
        return "CL"
    else:
        return "GB"


# 功能函数块结束------------------------------


# 主进程开始---------------------------------
if __name__ == '__main__':
    name = sys.argv[1:2]
    name=['汇总']
    if (len(name) < 1 or len(name[0]) < 1):
        print('RTC sheet name is empty ')
        sys.exit()
    else:
        # rtcSheetName = name[0]
        print('name is ', name[0])

sTime = time.time()
print("CSV文件转存xlsx文件......")
# 创建一个workbook
xlsbook = openpyxl.Workbook()
# 创建一个worksheet
newsheet = xlsbook.create_sheet('RTC Bug汇总', -1)
# newsheet.title='Open assigned to me'

with open(csvFile, 'r', encoding='utf-16') as csv:
    row = csv.read().split('"\n')
    for line in row:
        line = re.sub('"', '', line)
        cells = line.split('	')
        newsheet.append(cells)

extCol = newsheet.max_column+1
align = Alignment(horizontal='center', vertical='center')
for i in range(0, len(extName)):
    newsheet.cell(row=1, column=extCol+i).value = extName[i]
    newsheet.cell(row=1, column=extCol+i).alignment = align
    newsheet.cell(row=1, column=extCol +
                  i).fill = PatternFill(start_color=extendColcolor, fill_type="solid")

xlsbook.save(newFile)
print("文件转存完毕......")


print("读取表格文件数据")
newSheet = pd.read_excel(newFile, sheet_name='RTC Bug汇总')
rtcSheet = pd.read_excel(rtcFile, sheet_name=name[0])

# 删除"Type"列不是"Bug"的记录
newSheet = newSheet[newSheet['Type'] == 'Bug']

print("新增及已存在bug记录处理")

# 将newSheet和rtcSheet设置为a_id和b_id为索引
newSheet.set_index('Id', inplace=True)
rtcSheet.set_index('Id', inplace=True)
newSheet['签入日期']=newSheet['签入日期'].apply(pd.to_datetime)
for col in newSheet.columns:
    if newSheet[col].dtype!=rtcSheet[col].dtype:
        newSheet[col]=newSheet[col].astype(rtcSheet[col].dtype)
        # print(f"\n{col} 列类型不匹配: newSheet {newSheet[col].dtype}, rtcSheet {rtcSheet[col].dtype}")

# 使用merge函数合并两个数据表，on参数设置为'Id'，表示以'Id'列作为主键进行合并
merged = pd.merge(newSheet, rtcSheet, on='Id')

# 初始化一个空数组id_tmp
id_tmp = []


if not rtcSheet.index.is_unique:
    print("rtcSheet索引不唯一，存在重复值：")
    print(rtcSheet.index[rtcSheet.index.duplicated()])
    sys.exit()

if not newSheet.index.is_unique:
    print("newSheet索引不唯一，存在重复值：")
    print(newSheet.index[newSheet.index.duplicated()])
    sys.exit()

# 使用update函数更新newSheet中的记录
newSheet.update(rtcSheet[extName])

# 之前被标记已转出但依然挂着的记录标记为已打回
newSheet.loc[newSheet['研发状态开发更新']=='已转出', '研发状态开发更新'] = '已转回'
newSheet.loc[newSheet['研发状态开发更新']=='已转出', '研发状态开发更新'] = '已转回'

# 补充签入日期
newSheet.loc[newSheet['签入日期'].isnull(), '签入日期'] = datetime.now()

# 计算超期天数
newSheet['Creation Date'] = newSheet['Creation Date'].str.replace('上午', '')
newSheet['Creation Date'] = newSheet['Creation Date'].str.replace('下午', '')
# newSheet['Due Date']=datetime.strptime(newSheet['Due Date'],'%Y-%m-%d %p%I:%M')
newSheet['Creation Date'] = pd.to_datetime(newSheet['Creation Date'])
newSheet.loc[(newSheet['研发状态开发更新'] != '已转出') & (newSheet['Creation Date'].notnull(
)), '超期/天（Creation Date）'] = (datetime.now() - newSheet['Creation Date']).dt.days

# 使用concat函数将不存在于newSheet中的b_id记录追加到newSheet的末尾
print("追加已迁出记录")
rtcSheet.loc[~rtcSheet.index.isin(newSheet.index), '研发状态开发更新'] = '已转出'
newSheet = pd.concat([newSheet, rtcSheet[~rtcSheet.index.isin(newSheet.index)]])


# 补充bug所属车型平台
newSheet.loc[newSheet['分支车型'].isnull(), '分支车型'] = newSheet['Platform Vehicle Model'].apply(getPlatform)
# newSheet['车型'] = newSheet['Platform Vehicle Model'].apply(getCarType)

newSheet.to_excel(saveFile)
print("数据处理完毕，开始调整单元格式......")
xl_book = openpyxl.load_workbook(saveFile, read_only=False, data_only=True)
xl_sheet = xl_book['Sheet1']

rmv_fill = PatternFill(fill_type="solid", fgColor="808080")
addmust_fill = PatternFill(fill_type="solid", fgColor="ED4197")

# 遍历状态列
for row in xl_sheet.iter_rows():
    if (str(row[0]) in id_tmp):
        for c in row:
            c.fill = addmust_fill
    cell = row[13]
    # 如果状态列的值是“已转出”，就把当前行的背景色设置为灰色
    if (cell.value == '已转出'):
        for cell in row:
            cell.fill = rmv_fill

for col in ['M','P','Q','R','S','U']:
    for cell in xl_sheet[col]:
        cell.number_format = 'YYYY"年"M"月"D"日"'

xl_book.save(saveFile)
xl_book.close()

eTime = time.time()
print("数据处理完成，全程耗时：", str((eTime-sTime)*1000), "毫秒。")
