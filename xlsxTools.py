import openpyxl

sourceFile = "./DataFile/RTC list new.xlsx"
blankChars="#N/A"

data_book = openpyxl.load_workbook(sourceFile, read_only=False, data_only=True)
print(data_book.sheetnames)
data_sheet = data_book["Sheet1"]

maxRow=data_sheet.max_row
# print(maxRow)
# print(data_sheet[2][2].value)
# print(data_sheet[2][18].value)
# print(data_sheet[2][19].value)

i=2
while i<=maxRow:
    # 根据表格中"PlatformVehicleModel"-S列内容给"车型"-U列赋值
    # if (blankChars==data_sheet[i][20].value):
    #     if("E2-2" in data_sheet[i][18].value):
    #         data_sheet[i][20].value="E22"
    #     elif ("458" in data_sheet[i][18].value):
    #         data_sheet[i][20].value="458"
    #     elif ("B233" in data_sheet[i][18].value):
    #         data_sheet[i][20].value="B233"
    #     elif ("B223" in data_sheet[i][18].value):
    #         data_sheet[i][20].value="B223"
    #     else:
    #         data_sheet[i][20].value="E22"

    # 根据表格中"Summary"-C列描述给"模块"-T列为"#N/A"的赋值
    if (blankChars==data_sheet[i][19].value):
        if("[adas settings]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="ADAS_Settings"
        elif ("[adas]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="ADAS"
        elif ("[ambient_light]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Ambient_Light"
        elif ("[ambient light]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Ambient_Light"
        elif ("[audio]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Audio_Basic"
        elif ("[media]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Audio_Basic"
        elif ("[audio_basic]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Audio_Basic"
        elif ("[patac_seats]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="CabinMode"
        elif ("[climate]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Climate"
        elif ("[cluster]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Cluster"
        elif ("[cluster_adas]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Cluster_ADAS"
        elif ("[cluster_alert]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Cluster_Alert"
        elif ("[climate]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Climate"
        elif ("[cluster]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Cluster"
        elif ("[cluster_adas]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Cluster_ADAS"
        elif ("[cluster_alert]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Cluster_Alert"
        elif ("[cluster_guage]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Cluster_Guage"
        elif ("[cluster_warning]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Cluster_Warning"
        elif ("[cluster_zone2]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Cluster_Zone2"
        elif ("[drive_mode]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Drive_Mode"
        elif ("[driving_info]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Driving_Info"
        elif ("[easter_egg]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Easter_egg"
        elif ("[hud_adas]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="HUD_ADAS"
        elif ("[hud hmi]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="HUD_ADAS"
        elif ("[performance]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Performance"
        elif ("[setting]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Settings"
        elif ("[settings]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Settings"
        elif ("[vehicle_control]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Vehicle_Control"
        elif ("Vehiclecontrol" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Vehicle_Control"
        elif ("[Vehicle Control]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Vehicle_Control"
        elif ("vehicle control" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Vehicle_Control"
        elif ("[vehicle_info]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Vehicle_Info"
        elif ("[vehicle info]" in data_sheet[i][2].value.lower()):
            data_sheet[i][19].value="Vehicle_Info"
    i=i+1

data_book.save(sourceFile)
data_book.close()
print("Mission completed.")