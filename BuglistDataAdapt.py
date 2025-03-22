import datetime
import openpyxl
from openpyxl.styles import PatternFill
import time

yesterdayFile = "./DataFile/yesterday.xlsx"
todayFile = "./DataFile/today.xlsx"
saveFile = "./DataFile/newBuglist.xlsx"

extendColcolor = "FDE9D9"
outRecordcolor = "969696"

commitDateCol = 9
removeCol=10


def keyIsExistRtc(val):
    cell = None
    for c in yesterdaySheet['B']:
        if (c.row == 1):
            continue
        if (str(val) == str(c.value)):
            # print(val)
            cell = c
    return cell

def keyIsExistNew(val):
    cell = None
    for c in todaySheet['B']:
        if (c.row == 1):
            continue
        if (str(val) == str(c.value)):
            # print(val)
            cell = c
    return cell


sTime = time.time()
print("准备：读取表格文件数据")
today_book = openpyxl.load_workbook(todayFile, read_only=False, data_only=True)
todaySheet = today_book['today']
yesterday_book = openpyxl.load_workbook(yesterdayFile, read_only=False, data_only=True)
yesterdaySheet = yesterday_book['yesterday']

print("第一步：上期已存在记录备注信息迁移")
for id in todaySheet['B']:
    if (id.row == 1):
        continue
    # print(id.value,id.row,id.column)
    rtcCell = keyIsExistRtc(id.value)
    if (None != rtcCell):
        for i in range(1, todaySheet.max_column + 1):
            todaySheet.cell(row=id.row, column=i).fill = PatternFill(start_color=extendColcolor, fill_type="solid")
    else:
        todaySheet.cell(row=id.row, column=commitDateCol, value=datetime.datetime.today())

print("第二步：已迁出的记录尾部追加保存")
insNo = todaySheet.max_row+1
for id in yesterdaySheet['B']:
    if (id.row == 1):
        continue
    newCell = keyIsExistNew(id.value)
    if (None == newCell):
        tmpRow = yesterdaySheet[id.row]
        todaySheet.insert_rows(insNo)
        for i in range(1, todaySheet.max_column+1):
            if (i == removeCol):
                todaySheet.cell(insNo, i).value = datetime.datetime.today()
            else:
                todaySheet.cell(insNo, i).value = tmpRow[i-1].value
            todaySheet.cell(insNo, i).fill = PatternFill(start_color=outRecordcolor, fill_type="solid")
        insNo = insNo+1
print("信息追加完成")

today_book.save(saveFile)
today_book.close()
yesterday_book.close()
eTime = time.time()
print("数据处理完成，全程耗时：", str((eTime-sTime)*1000), "毫秒。")